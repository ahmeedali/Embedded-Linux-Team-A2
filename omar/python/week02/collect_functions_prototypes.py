#!/usr/bin/python3
from pyclibrary import CParser
import openpyxl

def write_excel_file(file_name, sheet_name, data):
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.create_sheet(title=sheet_name)
    
    # Write data to the excel sheet
    if data:
        for row in data:
            sheet.append(row)
    else:
        print("No data available")

    # Save and close excel file
    workbook.save(filename=(file_name + '.xlsx'))
    workbook.close()

def read_excel_file(file_path,sheet_name):

    try:
        # Load excel file using it's path
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        # Print the exception message
        print(f"An exception occurred: {str(e)}")

    try:
       # Activate required sheet to read from
        sheet = workbook.get_sheet_by_name(sheet_name)
    except Exception as e:
        # Print the exception message
        print(f"An exception occurred: {str(e)}")

    # Parse sheet data and print it out
    for row in sheet.iter_rows(values_only=True):
        print(row)

    # Close Excel file 
    workbook.close()


# Function to parse given C header file prototypes and return them in a dictionary
def parse_prototypes(file_path) -> dict:
    prototypes_dict = {}
    
    try:
        parser = CParser(file_path)
        parser.process_all()
        prototypes_dict = parser.defs['functions']
    except Exception as e:
        # Print the exception message
        print(f"An exception occurred: {str(e)}")

    return prototypes_dict

# function to format parsed functions' prototypes dictionary data into lists
def format_parsed_prototypes(parsed_dict) -> list:
    formatedList = []
    startIndex = 0

    for k, v in parsed_dict.items():
        startIndex += 1
        formatedList.append([str(k),str(v),'IDX0'+str(startIndex)]) 
    
    return formatedList

def main():
    # Read header file path from the user
    headerFilePath = input('Header file path: ')

    # Read excel file name to be created from the user
    excelFileName = input('New excel file name: ')

    # Read sheet name to be created from the user
    sheetName = input('Sheet Name: ')   
    
    # Call parse_prototypes to return a dictionary containing parsed prototypes
    prototypes_dict = parse_prototypes(headerFilePath)

    # Call format_parsed_prototypes to return a list containing formated prototypes
    prototypes_ls = format_parsed_prototypes(prototypes_dict)

    # Call write_excel_file to fill a new sheet in new excel file with the parsed formated prototypes
    write_excel_file(file_name=excelFileName, sheet_name=sheetName, data=prototypes_ls) 

    # Call read_excel_file to represent the filled data in the new sheet inside the newly created excel file
    read_excel_file(file_path=(excelFileName + '.xlsx'),sheet_name=sheetName)

if __name__ == '__main__':
    main()